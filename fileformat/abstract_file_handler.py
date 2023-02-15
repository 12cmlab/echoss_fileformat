import abc
import io
import pandas as pd

import csv
import json
import xml.etree.ElementTree as ET
import openpyxl


class AbstractFileFormatHandler(metaclass=abc.ABCMeta):
    @abc.abstractmethod
    def load(self, from_stream):
        pass

    @abc.abstractmethod
    def loads(self, from_string):
        pass

    @abc.abstractmethod
    def get(self, xpath):
        pass

    @abc.abstractmethod
    def update(self, xpath, new_data):
        pass

    @abc.abstractmethod
    def dump(self, to_stream):
        pass

    @abc.abstractmethod
    def dumps(self):
        pass


class JsonHandler(AbstractFileFormatHandler):
    def __init__(self):
        self.data = None

    def load(self, from_stream):
        self.data = json.load(from_stream)
        return self.data

    def loads(self, from_string):
        self.data = json.loads(from_string)
        return self.data

    def get(self, xpath):
        return self.data.get(xpath)

    def update(self, xpath, new_data):
        self.data[xpath] = new_data

    def dump(self, to_stream):
        json.dump(self.data, to_stream)

    def dumps(self):
        return json.dumps(self.data)


class XmlHandler(AbstractFileFormatHandler):
    def __init__(self):
        self.tree = None

    def load(self, from_stream):
        self.tree = ET.parse(from_stream)

    def loads(self, from_string):
        self.tree = ET.fromstring(from_string)

    def get(self, xpath):
        for element in self.tree.iter():
            if element.text == xpath:
                return element
        return None

    def update(self, xpath, new_data):
        element = self.get(xpath)
        if element is not None:
            element.text = new_data
        else:
            new_element = ET.Element(xpath)
            new_element.text = new_data
            self.tree.getroot().append(new_element)

    def dump(self, to_stream):
        self.tree.write(to_stream)

    def dumps(self):
        return ET.tostring(self.tree.getroot(), encoding="unicode")


class CsvHandler(AbstractFileFormatHandler):
    def __init__(self, **kwargs):
        self.data = []
        self.kwargs = kwargs
        self.header = []

    def load(self, from_stream):
        csv_reader = csv.reader(from_stream, **self.kwargs)
        # set self.header if header setting
        for row in csv_reader:
            self.data.append(row)
        return self.data

    def loads(self, from_string):
        stream = io.StringIO(from_string)
        return self.load(stream)

    def get(self, content):
        for row in self.data:
            if content in row:
                return row
        return None

    def update(self, content, new_data):
        row = self.get(content)
        if row is not None:
            index = self.data.index(row)
            self.data[index] = new_data
        else:
            self.data.append(new_data)

    def dump(self, to_stream):
        csv_writer = csv.writer(to_stream, **self.kwargs)
        for row in self.data:
            csv_writer.writerow(row)

    def dumps(self):
        stream = io.StringIO()
        csv_writer = csv.writer(stream)
        for row in self.data:
            csv_writer.writerow(row)
        return stream.getvalue()


class ExcelFileFormatHandler(AbstractFileFormatHandler):
        def __init__(self):
            self.data = pd.DataFrame()

        def load(self, stream):
            self.data = pd.read_excel(stream)

        def loads(self, string):
            file = io.BytesIO(string)
            self.data = pd.read_excel(file)

        def get(self, content):
            for row in self.data.iterrows():
                if content in row:
                    return row
            return None

        def update(self, content, new_data):
            index = None
            for i, row in self.data.iterrows():
                if content in row:
                    index = i
                    break
            if index is not None:
                self.data.loc[index] = new_data
            else:
                self.data = self.data.append(new_data, ignore_index=True)

        def dump(self, stream):
            self.data.to_excel(stream, index=False)

        def dumps(self):
            file = io.BytesIO()
            self.data.to_excel(file, index=False)
            return file.getvalue()


class ExcelFileFormatHandler2(AbstractFileFormatHandler):
    def __init__(self):
        self.workbook = None

    def load(self, stream):
        self.workbook = openpyxl.load_workbook(stream)

    def loads(self, string):
        # Not supported in openpyxl
        raise NotImplementedError("ExcelFileFormatHandler does not support loads")

    def get(self, path):
        sheet_name, row_index, column_index = path
        sheet = self.workbook[sheet_name]
        return sheet.cell(row=row_index, column=column_index).value

    def update(self, path, value):
        sheet_name, row_index, column_index = path
        sheet = self.workbook[sheet_name]
        sheet.cell(row=row_index, column=column_index).value = value

    def dump(self, stream):
        self.workbook.save(stream)

    def dumps(self):
        # Not supported in openpyxl
        raise NotImplementedError("ExcelFileFormatHandler does not support dumps")